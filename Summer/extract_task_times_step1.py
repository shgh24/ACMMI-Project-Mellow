import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

def parse_jsonl(file_path):
    """Parse a JSONL file and return list of entries"""
    entries = []
    with open(file_path, 'r') as f:
        for line in f:
            if line.strip():
                entries.append(json.loads(line))
    return entries

def format_duration(seconds):
    """Format seconds into readable duration"""
    return str(timedelta(seconds=int(seconds)))

def extract_task_times(jsonl_file):
    """
    Extract start and end times for each task from the JSONL file.
    
    Tasks identified:
    - Descriptive Task: TASK_STARTED/DESCRIPTIVE_TASK_STARTED -> DESCRIPTIVE_COUNTDOWN_AUTO_TRANSITION
    - Stroop Task: STROOP_TASK_STARTED -> STROOP_VIDEO_END_TRANSITION
    - Math Task: MATH_TASK_STARTED -> MATH_COUNTDOWN_AUTO_TRANSITION
    - Content Performance Task: CONTENT_PERFORMANCE_SCREEN_DISPLAYED -> CONTENT_PERFORMANCE_COMPLETED
    """
    
    # Load actions from JSONL file
    actions = parse_jsonl(jsonl_file)
    
    # Extract participant ID (should be the same for all actions in a file)
    participant_id = None
    if actions and 'participant_id' in actions[0]:
        participant_id = actions[0]['participant_id']
    
    # Define task boundaries
    task_boundaries = {
        'Pre-Study Relaxation': {
            'start_events': ['RELAXATION_COUNTDOWN_STARTED'],
            'end_events': ['RELAXATION_COUNTDOWN_AUTO_TRANSITION']
        },
        'Descriptive Task': {
            'start_events': ['TASK_STARTED', 'DESCRIPTIVE_TASK_STARTED'],
            'end_events': ['DESCRIPTIVE_COUNTDOWN_AUTO_TRANSITION']
        },
        'Stroop Task': {
            'start_events': ['STROOP_VIDEO_STARTED', 'STROOP_VIDEO_STARTED_3_MIN'],  # First video start
            'end_events': ['STROOP_VIDEO_END_TRANSITION']  # End of second video
        },
        'Math Task': {
            'start_events': ['MATH_TASK_STARTED'],
            'end_events': ['MATH_COUNTDOWN_AUTO_TRANSITION']
        },
        'Content Performance Task': {
            'start_events': ['CONTENT_PERFORMANCE_SCREEN_DISPLAYED'],
            'end_events': ['CONTENT_PERFORMANCE_COMPLETED']
        },
        'Post-Study Relaxation': {
            'start_events': ['POST_STUDY_COUNTDOWN_STARTED'],
            'end_events': ['POST_STUDY_COUNTDOWN_AUTO_TRANSITION']
        }
    }
    
    # Extract task information
    tasks = []
    
    for task_name, boundaries in task_boundaries.items():
        start_action = None
        end_action = None
        
        # Special handling for Stroop Task - find first video start
        if task_name == 'Stroop Task':
            # Find the first STROOP_VIDEO_STARTED event (any variant)
            for action in actions:
                if action['action_type'].startswith('STROOP_VIDEO_STARTED'):
                    start_action = action
                    break
            
            # Find the end event (STROOP_VIDEO_END_TRANSITION) after the start
            if start_action:
                start_idx = actions.index(start_action)
                for action in actions[start_idx:]:
                    if action['action_type'] in boundaries['end_events']:
                        end_action = action
                        break
        else:
            # Standard handling for other tasks
            # Find start event
            for action in actions:
                if action['action_type'] in boundaries['start_events']:
                    start_action = action
                    break
            
            # Find end event (after start)
            if start_action:
                start_idx = actions.index(start_action)
                for action in actions[start_idx:]:
                    if action['action_type'] in boundaries['end_events']:
                        end_action = action
                        break
        
        if start_action and end_action:
            # Calculate duration
            duration_seconds = end_action['session_duration_seconds'] - start_action['session_duration_seconds']
            
            # Extract task details
            task_info = {
                'participant_id': participant_id,
                'task_name': task_name,
                'start_time': {
                    'local': start_action['timestamp']['local'],
                    'utc': start_action['timestamp']['utc'],
                    'unix': start_action['timestamp']['unix'],
                    'session_duration_seconds': start_action['session_duration_seconds']
                },
                'end_time': {
                    'local': end_action['timestamp']['local'],
                    'utc': end_action['timestamp']['utc'],
                    'unix': end_action['timestamp']['unix'],
                    'session_duration_seconds': end_action['session_duration_seconds']
                },
                'duration_seconds': duration_seconds,
                'duration_formatted': format_duration(duration_seconds),
                'start_action_type': start_action['action_type'],
                'end_action_type': end_action['action_type'],
                'start_details': start_action.get('details', ''),
                'end_details': end_action.get('details', '')
            }
            
            # Add task-specific information
            if task_name == 'Content Performance Task':
                # Extract task type from details if available
                details = start_action.get('details', '')
                if 'Task:' in details:
                    task_info['task_type'] = details.split('Task:')[1].strip()
                elif 'task' in start_action.get('details', '').lower():
                    task_info['task_type'] = start_action.get('details', '')
            
            tasks.append(task_info)
    
    return tasks

def print_task_times(tasks):
    """Print task times in a readable format"""
    print("=" * 80)
    print("TASK START AND END TIMES")
    print("=" * 80)
    print()
    
    if not tasks:
        print("No tasks found in the log file.")
        return
    
    for i, task in enumerate(tasks, 1):
        print(f"{i}. {task['task_name']}")
        if task.get('participant_id'):
            print(f"   Participant ID: {task['participant_id']}")
        print(f"   Start: {task['start_time']['local']}")
        print(f"   End:   {task['end_time']['local']}")
        print(f"   Duration: {task['duration_formatted']} ({task['duration_seconds']:.2f} seconds)")
        print(f"   Start Event: {task['start_action_type']}")
        print(f"   End Event: {task['end_action_type']}")
        if task.get('task_type'):
            print(f"   Task Type: {task['task_type']}")
        if task.get('start_details'):
            print(f"   Start Details: {task['start_details']}")
        print()

def save_task_times_json(tasks, output_file):
    """Save task times to a JSON file"""
    with open(output_file, 'w') as f:
        json.dump(tasks, f, indent=2)
    print(f"Task times saved to {output_file}")

def save_task_times_csv(tasks, output_file):
    """Save task times to a CSV file"""
    import csv
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        # Header
        writer.writerow([
            'Participant ID', 'Task Name', 'Start Time (Local)', 'End Time (Local)',
            'Duration (seconds)', 'Duration (formatted)',
            'Start Action Type', 'End Action Type', 'Task Type', 'Start Details'
        ])
        
        # Data rows
        for task in tasks:
            writer.writerow([
                task.get('participant_id', ''),
                task['task_name'],
                task['start_time']['local'],
                task['end_time']['local'],
                task['duration_seconds'],
                task['duration_formatted'],
                task['start_action_type'],
                task['end_action_type'],
                task.get('task_type', ''),
                task.get('start_details', '')
            ])
    
    print(f"Task times saved to {output_file}")

def save_task_times_excel(tasks, output_file):
    """Save task times to an Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Error: openpyxl is required for Excel export. Install it with: pip install openpyxl")
        return
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task Times"
    
    # Define header
    headers = [
        'Participant ID', 'Task Name', 'Start Time (Local)', 'End Time (Local)',
        'Duration (seconds)', 'Duration (formatted)',
        'Start Action Type', 'End Action Type', 'Task Type', 'Start Details'
    ]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task.get('participant_id', ''))
        ws.cell(row=row_num, column=2, value=task['task_name'])
        ws.cell(row=row_num, column=3, value=task['start_time']['local'])
        ws.cell(row=row_num, column=4, value=task['end_time']['local'])
        ws.cell(row=row_num, column=5, value=task['duration_seconds'])
        ws.cell(row=row_num, column=6, value=task['duration_formatted'])
        ws.cell(row=row_num, column=7, value=task['start_action_type'])
        ws.cell(row=row_num, column=8, value=task['end_action_type'])
        ws.cell(row=row_num, column=9, value=task.get('task_type', ''))
        ws.cell(row=row_num, column=10, value=task.get('start_details', ''))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"Task times saved to {output_file}")

def process_directory(log_dir, output_dir=None):
    """Process all action files in a directory and combine results"""
    log_path = Path(log_dir)
    if not log_path.exists():
        print(f"Error: Directory not found: {log_path}")
        return []
    
    all_tasks = []
    
    # Find all action files
    action_files = sorted(log_path.glob('actions_*.jsonl'))
    
    if not action_files:
        print(f"No action files found in {log_path}")
        return []
    
    print(f"Found {len(action_files)} action file(s) in {log_path}")
    print()
    
    for action_file in action_files:
        print(f"Processing: {action_file.name}")
        tasks = extract_task_times(action_file)
        all_tasks.extend(tasks)
        print(f"  Found {len(tasks)} tasks")
    
    return all_tasks

def process_multiple_participants(logs_base_dir, participant_ids=None):
    """Process multiple participant directories"""
    base_path = Path(logs_base_dir)
    all_tasks = []
    
    # If participant_ids not specified, discover all participant directories
    if participant_ids is None:
        # Find all directories in logs_base_dir that look like participant IDs
        participant_dirs = [d for d in base_path.iterdir() 
                           if d.is_dir() and (d.name.startswith('MIT') or d.name.startswith('KITTY'))]
        participant_ids = sorted([d.name for d in participant_dirs])
        print(f"Auto-discovered {len(participant_ids)} participant directories: {', '.join(participant_ids)}")
    
    for participant_id in participant_ids:
        participant_dir = base_path / participant_id
        if participant_dir.exists():
            print(f"\n{'='*80}")
            print(f"Processing {participant_id}")
            print(f"{'='*80}")
            tasks = process_directory(participant_dir)
            all_tasks.extend(tasks)
        else:
            print(f"Warning: Directory not found for {participant_id}: {participant_dir}")
    
    return all_tasks

def main():
    parser = argparse.ArgumentParser(
        description='Extract start and end times for each task from JSONL log file(s)'
    )
    # Default logs path
    default_logs_path = Path('/Users/summerghorbani/Documents/MIT_projects/Molly/logs')
    
    parser.add_argument('input_path', nargs='?', 
                       default=str(default_logs_path) if default_logs_path.exists() else 'actions_20251016_174631.jsonl',
                       help='Path to JSONL file or directory containing log files (default: /Users/summerghorbani/Documents/MIT_projects/Molly/logs)')
    parser.add_argument('--json', '-j', 
                       help='Save output to JSON file')
    parser.add_argument('--csv', '-c',
                       help='Save output to CSV file')
    parser.add_argument('--excel', '-e', '--xlsx', '-x',
                       help='Save output to Excel file (.xlsx)')
    parser.add_argument('--directory', '-d', action='store_true',
                       help='Treat input_path as a directory and process all action files')
    parser.add_argument('--participants', '-p', nargs='+',
                       help='Process specific participant directories (e.g., MIT001 MIT002). If not specified and input_path is logs directory, all participants will be processed.')
    
    # Default output directory: Molly/Results/Tasks
    script_dir = Path(__file__).parent
    default_output = script_dir.parent.parent.parent / 'Results' / 'Tasks'
    parser.add_argument('--output-dir', '-o', type=str, default=str(default_output),
                       help=f'Output directory for results (default: {default_output})')
    parser.add_argument('--all', '-a', action='store_true',
                       help='Process all participants in the logs directory (auto-discover)')
    
    args = parser.parse_args()
    
    input_path = Path(args.input_path)
    
    if not input_path.exists():
        print(f"Error: Path not found: {input_path}")
        return
    
    # Process multiple participants if specified or if --all flag is used
    if args.participants or args.all or (input_path.name == 'logs' and input_path.is_dir()):
        # Assume input_path is the base logs directory
        participant_list = args.participants if args.participants else None
        tasks = process_multiple_participants(input_path, participant_list)
    # Determine if we're processing a directory or single file
    elif args.directory or input_path.is_dir():
        tasks = process_directory(input_path)
    else:
        # Process single file
        tasks = extract_task_times(input_path)
        print_task_times(tasks)
    
    if not tasks:
        print("No tasks found.")
        return
    
    # Create output directory if it doesn't exist
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"\nOutput directory: {output_dir.absolute()}")
    
    # Default to Excel output if no format specified and processing directory
    if (args.directory or input_path.is_dir()) and not any([args.json, args.csv, args.excel]):
        # Generate default Excel filename
        default_excel = f"{input_path.name}_task_times.xlsx" if input_path.is_dir() else "task_times.xlsx"
        output_path = output_dir / default_excel
        save_task_times_excel(tasks, output_path)
        print(f"\nSaved to {output_path} (default)")
    
    # Save to files if requested
    if args.json:
        output_path = output_dir / args.json if not Path(args.json).is_absolute() else Path(args.json)
        save_task_times_json(tasks, output_path)
    
    if args.csv:
        output_path = output_dir / args.csv if not Path(args.csv).is_absolute() else Path(args.csv)
        save_task_times_csv(tasks, output_path)
    
    if args.excel:
        output_path = output_dir / args.excel if not Path(args.excel).is_absolute() else Path(args.excel)
        save_task_times_excel(tasks, output_path)
    
    # Print summary
    if tasks:
        total_duration = sum(task['duration_seconds'] for task in tasks)
        print("\n" + "=" * 80)
        print("SUMMARY")
        print("=" * 80)
        print(f"Total tasks found: {len(tasks)}")
        print(f"Total task duration: {format_duration(total_duration)} ({total_duration:.2f} seconds)")
        
        # Group by participant
        from collections import defaultdict
        by_participant = defaultdict(list)
        for task in tasks:
            by_participant[task.get('participant_id', 'Unknown')].append(task)
        
        print(f"\nTasks by participant:")
        for participant_id, participant_tasks in sorted(by_participant.items()):
            print(f"  {participant_id}: {len(participant_tasks)} tasks")

if __name__ == "__main__":
    main()

